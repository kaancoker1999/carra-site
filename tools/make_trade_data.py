#!/usr/bin/env python3
"""Build assets/trade-data.json for the LUMIA trade area.

Prices come from the pricing workbook (PRICING_XLSX below). The workbook holds
GROUP 1 prices; groups 2-4 are derived with GROUP_MULT. Dollar figures scale
with the multiplier, percentage modifiers do not.

Dealers get an access code. Each code decrypts exactly one customer group's
price list. Group membership is never exposed in the UI, and price data sits
encrypted in the repo (AES-256-GCM, key derived from the code via PBKDF2).

Real dealer codes live in tools/dealers.local.json (gitignored — NEVER commit
codes to the public repo): {"LUMIA-XXXX": 1, ...}. Demo codes are the fallback.

Re-run after changing the workbook or dealer list:
    python3 tools/make_trade_data.py
"""
import base64
import hashlib
import json
import os
import pathlib

import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

HERE = pathlib.Path(__file__).parent
OUT = HERE.parent / "assets" / "trade-data.json"
PRICING_XLSX = HERE / "pricing.local.xlsx"  # gitignored — base (group 1) prices
DRAPERY_XLSX = HERE / "drapery-pricing.local.xlsx"  # gitignored — drapery grids
ITER = 150_000

GROUP_MULT = [1.00, 1.15, 1.30, 1.50]  # group 1 = workbook prices

DEALERS = {
    "LUMIA-DEMO-G1": 1,
    "LUMIA-DEMO-G2": 2,
    "LUMIA-DEMO-G3": 3,
    "LUMIA-DEMO-G4": 4,
}
_local = HERE / "dealers.local.json"
if _local.exists():
    DEALERS = {k: int(v) for k, v in json.loads(_local.read_text()).items()}


# ── workbook parsing ────────────────────────────────────────────────────────

def sheet_rows(ws):
    out = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        while vals and vals[-1] is None:
            vals.pop()
        out.append(vals)
    return out


def read_grid(rows, header_row, n_rows):
    """header_row: 1-based row index of the widths header (first cell None)."""
    header = rows[header_row - 1]
    widths = [int(v) for v in header[1:]]
    body = []
    for r in rows[header_row:header_row - 1 + n_rows + 1]:
        if not r or r[0] is None:
            break
        body.append({"label": f"{int(r[0])}″", "vals": [round(float(v), 2) for v in r[1:1 + len(widths)]]})
    return {"cols": [f"{w}″" for w in widths], "corner": "H ▾ / W ▸", "rows": body}


def find_headers(rows):
    """Rows that look like a width header: first cell empty, rest ints."""
    hits = []
    for i, r in enumerate(rows, 1):
        if len(r) >= 5 and r[0] is None and all(isinstance(v, (int, float)) for v in r[1:]):
            hits.append(i)
    return hits


def parse_workbook():
    wb = openpyxl.load_workbook(PRICING_XLSX, data_only=True)

    # ── Cellular: stacked grids under LF / BlackOut, then surcharges ──
    cel = sheet_rows(wb["Cellular"])
    cel_titles = [
        "Light Filtering · Single Cell",
        "Light Filtering · Double Cell",
        "Light Filtering · Single Cell 3/4″",
        "Light Filtering · Triple Cell 3/8″",
        "Designer Prints · Single Cell 3/4″",
        "Sheer Woven · Single Cell 3/4″",
        "Blackout · Single Cell",
        "Blackout · Double Cell",
        "Blackout · Single Cell 3/4″",
    ]
    cel_headers = find_headers(cel)
    assert len(cel_headers) == len(cel_titles), f"Cellular: {len(cel_headers)} grids, expected {len(cel_titles)}"
    cel_grids = [dict(read_grid(cel, h, 40), name=t) for h, t in zip(cel_headers, cel_titles)]

    cellular_extras = [
        {"label": "Cordless", "note": "standard"},
        {"label": "Cordless TDBU", "usd": 30.5},
        {"label": "Two-on-one cordless", "usd": 28},
        {"label": "Two-on-one cordless TDBU", "usd": 80},
        {"label": "Continuous cord loop (CCL)", "usd": 10},
        {"label": "CCL + TDBU", "usd": 30.5},
        {"label": "Motorization", "usd": 136},
        {"label": "Motor + TDBU", "usd": 275},
        {"label": "10 ft USB cable", "usd": 9},
        {"label": "Remote control — 1 channel", "usd": 20},
        {"label": "Remote control — 5 channel", "usd": 33},
        {"label": "Remote control — 15 channel", "usd": 44},
        {"label": "Motionblinds WiFi hub", "usd": 160},
    ]

    # ── Roman: base grid expanded into one grid per fabric colour group ──
    rom = sheet_rows(wb["Roman"])
    rom_base = read_grid(rom, find_headers(rom)[0], 20)
    COLOUR_GROUPS = [("Colour group 1", 1.00), ("Colour group 2", 1.30),
                     ("Colour group 3", 1.55), ("Colour group 4", 1.80)]
    rom_grids = []
    for label, cg in COLOUR_GROUPS:
        g = json.loads(json.dumps(rom_base))
        for r in g["rows"]:
            r["vals"] = [round(v * cg, 2) for v in r["vals"]]
        g["name"] = label
        rom_grids.append(g)
    roman_extras = [
        {"label": "Seamless flat fold", "pct": 0},
        {"label": "Classic flat fold", "pct": 10},
        {"label": "Relaxed fold", "pct": 10},
        {"label": "Soft hobbled fold", "pct": 25},
        {"label": "White liner", "pct": 10},
        {"label": "Blackout liner", "pct": 20},
        {"label": "Edge banding", "pct": 50},
        {"label": "Continuous cord loop (CCL)", "usd": 15},
        {"label": "Cordless TDBU", "usd": 35},
        {"label": "CCL TDBU", "usd": 35},
        {"label": "Motorization", "usd": 136},
        {"label": "Motor + TDBU", "usd": 275},
        {"label": "10 ft USB cable", "usd": 10},
        {"label": "Remote control — 1 channel", "usd": 20},
        {"label": "Remote control — 5 channel", "usd": 30},
        {"label": "Remote control — 15 channel", "usd": 40},
        {"label": "Motionblinds WiFi hub", "usd": 145},
    ]
    roman_notes = [
        "Pick the grid for the fabric's colour group; fold and liner percentages apply to that grid price and combine.",
        "Secondary (seasonal) fabric available at 50% discount off the original price — the less expensive fabric is discounted.",
    ]

    # ── Pleated: two charts + surcharges ──
    ple = sheet_rows(wb["Pleated"])
    ple_headers = find_headers(ple)
    # first header row on this sheet belongs to the surcharge block header? No —
    # surcharge block has text first cells, so headers found are the two charts.
    ple_grids = [
        dict(read_grid(ple, ple_headers[0], 20), name="Chart A · Smooth"),
        dict(read_grid(ple, ple_headers[1], 20), name="Chart B · Crushed"),
    ]
    pleated_extras = [
        {"label": "Cordless", "note": "standard"},
        {"label": "Cordless TDBU", "usd": 23},
        {"label": "Two-on-one cordless", "usd": 50},
        {"label": "Two-on-one cordless TDBU", "usd": 85},
        {"label": "Continuous cord loop (CCL)", "usd": 10},
        {"label": "Motorization", "usd": 125},
        {"label": "Motor + TDBU", "usd": 250},
        {"label": "10 ft USB cable", "usd": 8},
        {"label": "Remote control — 1 channel", "usd": 18},
        {"label": "Remote control — 5 channel", "usd": 30},
        {"label": "Remote control — 15 channel", "usd": 40},
        {"label": "Motionblinds WiFi hub", "usd": 145},
    ]

    # ── Cell arches: cell-type rows x width cols, fixed height ──
    arc = sheet_rows(wb["Cell arches"])
    arc_headers = [i for i, r in enumerate(arc, 1) if len(r) > 2 and r[1] == "Height"]

    def arch_grid(hrow, name, n):
        header = arc[hrow - 1]
        widths = [int(v) for v in header[2:]]
        body = []
        for r in arc[hrow:hrow + n]:
            if not r or not isinstance(r[0], str):
                break
            body.append({"label": r[0].replace("Dbl", "Double"),
                         "vals": [round(float(v), 2) for v in r[2:2 + len(widths)]]})
        return {"name": name, "cols": [f"{w}″" for w in widths],
                "corner": "Cell / W ▸", "rows": body}

    arch_grids = [
        arch_grid(arc_headers[0], "Light Filtering", 3),
        arch_grid(arc_headers[1], "Blackout", 2),
    ]

    products = [
        {"id": "cellular", "name": "Cellular Shades", "grids": cel_grids,
         "extras": cellular_extras, "notes": []},
        {"id": "roman", "name": "Roman Shades", "grids": rom_grids,
         "extras": roman_extras, "notes": roman_notes},
        {"id": "pleated", "name": "Pleated Shades", "grids": ple_grids,
         "extras": pleated_extras, "notes": []},
        {"id": "arches", "name": "Cellular Arches", "grids": arch_grids,
         "extras": [{"label": "Oversize shipping (width over 42″)", "usd": 16.35}],
         "notes": ["Arch heights up to 50″."]},
        parse_drapery(),
    ]
    return products



# ── Drapery: 5 heading sheets x fabric groups A-D + liners + hardware ──────
DRAPERY_HEADINGS = {
    "WAVE FOLD": "Wave fold",
    "SINGLE PINCH PLEAT": "Single pinch pleat",
    "DOUBLE PINCH PLEAT": "Double pinch pleat",
    "TRIPLE PINCH PLEAT": "Triple pinch pleat",
    "EURO PLEAT": "Euro pleat",
}


def read_wgrid(rows, header_row):
    """Width-header grid where the header row carries quoted inch labels."""
    header = rows[header_row - 1]
    widths = [str(v).strip() for v in header[1:] if v is not None]
    body = []
    for r in rows[header_row:]:
        if not r or r[0] is None or not str(r[0]).strip().rstrip('"').isdigit():
            break
        body.append({"label": str(r[0]).strip().replace('"', "″"),
                     "vals": [round(float(v), 2) for v in r[1:1 + len(widths)]]})
    return {"cols": [w.replace('"', "″") for w in widths], "corner": "H ▾ / W ▸", "rows": body}


def parse_drapery():
    wb = openpyxl.load_workbook(DRAPERY_XLSX, data_only=True)
    grids = []
    for sheet, heading in DRAPERY_HEADINGS.items():
        rows = sheet_rows(wb[sheet])
        for i, r in enumerate(rows, 1):
            if r and isinstance(r[0], str) and r[0].strip().startswith("Group"):
                letter = r[0].strip().split()[1]
                grids.append(dict(read_wgrid(rows, i + 1), name=f"{heading} · Group {letter}"))
    # liners + hardware are identical on every sheet — take the first
    rows = sheet_rows(wb["WAVE FOLD"])
    for i, r in enumerate(rows, 1):
        if r and isinstance(r[0], str):
            t = r[0].strip()
            if t == "WHITE LINER":
                grids.append(dict(read_wgrid(rows, i + 1), name="Liner · White"))
            elif t == "BLACKOUT LINER":
                grids.append(dict(read_wgrid(rows, i + 1), name="Liner · Blackout"))
            elif t.startswith("ROD & TRACK"):
                header = rows[i]
                widths = [str(v).strip().replace('"', "″") for v in header[1:] if v is not None]
                body = []
                for hw in rows[i + 1:i + 3]:
                    body.append({"label": str(hw[0]).strip().title(),
                                 "vals": [round(float(v), 2) for v in hw[1:1 + len(widths)]]})
                grids.append({"name": "Hardware · Rod & Track", "cols": widths,
                              "corner": "/ W ▸", "rows": body})
    return {
        "id": "drapery", "name": "Drapery", "grids": grids, "extras": [],
        "notes": [
            "Prices per panel. Pair (centre split): each panel is half the entered width — the order form prices it as 2 × the half-width panel.",
            "Liner grids are added on top of the panel price at the same size.",
            "Track / rod priced by full width from the hardware table.",
        ],
    }


def scaled(products, mult):
    out = json.loads(json.dumps(products))  # deep copy
    for p in out:
        for g in p["grids"]:
            for r in g["rows"]:
                r["vals"] = [round(v * mult, 2) for v in r["vals"]]
        for e in p["extras"]:
            if "usd" in e:
                e["usd"] = round(e["usd"] * mult, 2)
    return out


def prices_for_group(gi, products):
    return {
        "currency": "USD",
        "products": scaled(products, GROUP_MULT[gi - 1]),
        "notes": ["Prices per unit in USD. Volume pricing available — contact us."],
    }


# ── crypto ──────────────────────────────────────────────────────────────────

def b64(b):
    return base64.b64encode(b).decode()


def derive(secret, salt):
    return hashlib.pbkdf2_hmac("sha256", secret.encode(), salt, ITER, dklen=32)


def encrypt(secret, plaintext):
    salt, iv = os.urandom(16), os.urandom(12)
    ct = AESGCM(derive(secret, salt)).encrypt(iv, plaintext, None)
    return {"s": b64(salt), "i": b64(iv), "c": b64(ct)}


def main():
    products = parse_workbook()
    n_groups = max(DEALERS.values())
    group_keys = [os.urandom(24).hex() for _ in range(n_groups)]

    groups = [
        encrypt(group_keys[gi], json.dumps(prices_for_group(gi + 1, products), separators=(",", ":")).encode())
        for gi in range(n_groups)
    ]
    dealers = [encrypt(code, group_keys[grp - 1].encode()) for code, grp in DEALERS.items()]
    dealers.sort(key=lambda d: d["s"])

    OUT.write_text(json.dumps({"kdf": {"iterations": ITER}, "dealers": dealers, "groups": groups}, indent=1))
    print(f"wrote {OUT} ({OUT.stat().st_size // 1024} KB): {len(dealers)} dealer codes, {n_groups} groups")


if __name__ == "__main__":
    main()
